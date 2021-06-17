import os
import io
from flask import Flask, request, make_response, render_template, abort, jsonify
from celery import chain
import sqlalchemy
from publisher.db import getEngine,db_execute,db_fetchone,db_fetchall,db_launch,db_start,db_end,db_error
import glob
import logging
import random
import time
import numpy
import datetime
import requests
import json
import copy
import urllib.request
from publisher.celery import tasks
from flask import current_app as app
import openpyxl
import smtplib
"""
#app = Flask(__name__)
#app.config['PROPAGATE_EXCEPTIONS'] = True
logging.basicConfig(filename='publisher.log',
	format='%(levelname)s:%(asctime)s in %(module)s: %(message)s',
	level=logging.INFO)
"""
###################################################
def loadConfig(sat):
	configfile = '{}.json'.format(sat)
	if os.path.exists(configfile):
		logging.warning(configfile+' was found in '+os.getcwd())
		with open(configfile) as infile:
			config = json.loads(infile.read())
		return config
	else:
		logging.warning(configfile+' was not found in '+os.getcwd())
		return None


###################################################
def c2jyd(calendardate):
	dt = numpy.datetime64(calendardate)
	year = calendardate[:4]
	jday = (dt - numpy.datetime64(year+'-01-01')).astype(int) + 1
	juliandate = year+'03d'.format(jday)
	return (juliandate,int(year),jday)

###################################################
def decodeDRDFilename(drdfilename):
# CBERS_4_AWFI_DRD_2017_04_22.15_05_00_CB11
# LANDSAT1_MSS_19750907.130000 
# CBERS2B_CCD_20100301.130915
# CBERS2B_WFI_20100304.125714
# AMAZONIA_1_WFI_DRD_2019_10_29.14_02_49
	filename = os.path.basename(drdfilename)
	sat = inst = date = name = number = calendardate = year = jday = None
	parts = filename.split('_')
	if len(parts) == 1:
		return (sat,inst,date,calendardate,year,jday)

	if (parts[0].find('CBERS') == 0 and parts[1].find('4') == 0) or parts[0].find('AMAZONIA') == 0 :
#		if len(parts) > 11:
#			return (sat,inst,date,calendardate,year,jday)
		sat = parts[0]+parts[1]
		inst = parts[2]
		calendardate = parts[4]+'-'+parts[5]+'-'+parts[6].split('.')[0]
		juliandate,year,jday = c2jyd(calendardate)
		date = parts[4]+'_'+parts[5]
	elif parts[0].find('LANDSAT') == 0 or parts[0].find('CBERS') == 0:
		if len(parts) < 3:
			return (sat,inst,date,calendardate,year,jday)
		sat = parts[0]
		inst = parts[1]
		calendardate = parts[2][0:4]+'-'+parts[2][4:6]+'-'+parts[2][6:8]
		juliandate,year,jday = c2jyd(calendardate)
		date = parts[2][0:4]+'_'+parts[2][4:6]
	logging.warning('decodeDRDFilename sat {} inst {} date {} calendardate {} year {} jday {}'.format(sat,inst,date,calendardate,year,jday))
	return (sat,inst,date,calendardate,year,jday)

#########################################
def buildBasicActivity(basicactivity,basedir = '/TIFF'):
	global config

# Get which satellite are in /TIFF
	satsdir = glob.glob('{}/*'.format(basedir))
	satsdict = {}
	for satdir in satsdir:
		satsdict[os.path.basename(satdir)] = satdir
	sats = [os.path.basename(x) for x in satsdir]
	basicactivity['message'] = ''
	basicactivity['sat'] = request.args.get('sat', None)
	if basicactivity['sat'] is None:
		basicactivity['message'] = 'Error - Please enter a satellite (sat).\nKnown satellites are:\n{}\n'.format(sats)
		return 1
	if basicactivity['sat'] not in satsdict:
		basicactivity['message'] = 'Error - Satellite {} is not known.\nKnown satellites are:\n{}\n'.format(basicactivity['sat'],sats)
		return 1
	config = loadConfig(basicactivity['sat'])
	if config is None:
		basicactivity['message'] = 'Error - Satellite {} has not a config file\n'.format(basicactivity['sat'])
		return 1
	basicactivity['inst'] = request.args.get('inst', None)
	if basicactivity['inst'] is not None and basicactivity['inst'] not in config['publish']['DN']:
		basicactivity['message'] = 'Error - Satellite {} has not a instrument {}\n'.format(basicactivity['sat'],basicactivity['inst'])
		return 1
# Manage the dates
# If start_date and end_date are specified, keep them if they are different.
	if request.args.get('start_date', None) is not None: 
		basicactivity['start_date'] = request.args.get('start_date', None)
	if request.args.get('end_date', None) is not None: 
		basicactivity['end_date'] = request.args.get('end_date', None)
# If they are the same, transform them to date
	if 'start_date' in basicactivity and 'end_date' in basicactivity and basicactivity['start_date'] == basicactivity['end_date']:
		basicactivity['date'] = basicactivity['start_date']
		del basicactivity['start_date']
		del basicactivity['end_date']
# If only one of them is specified, transform it to date
	if 'start_date' in basicactivity and 'end_date' not in basicactivity:
		#basicactivity['date'] = basicactivity['start_date']
		del basicactivity['start_date']
	if 'end_date' in basicactivity and 'start_date' not in basicactivity:
		#basicactivity['date'] = basicactivity['end_date']
		del basicactivity['end_date']
# If start_date and end_date have the same yyyy-mm, set the yyyy_mm field to speed up traversing the directories
	if 'start_date' in basicactivity and 'end_date' in basicactivity:
		sym = basicactivity['start_date'][:7].replace('-','_')
		eym = basicactivity['end_date'][:7].replace('-','_')
		if sym == eym:
			basicactivity['ym'] = sym
# if date is specified, set the yyyy_mm field to speed up traversing the directories
	if 'date' in basicactivity:
		basicactivity['ym'] = basicactivity['date'][:7].replace('-','_')

	#basicactivity['date'] = request.args.get('date', None)
	#basicactivity['ym'] = request.args.get('ym', None)
	basicactivity['level'] = request.args.get('level', None)
	basicactivity['drd'] = request.args.get('drd', None)
	basicactivity['rp'] = request.args.get('rp', None)
	basicactivity['path'] = request.args.get('path', None)
	#if basicactivity['path'] is not None: basicactivity['path'] = int(basicactivity['path'])
	basicactivity['row'] = request.args.get('row', None)
	#if basicactivity['row'] is not None: basicactivity['row'] = int(basicactivity['row'])
	basicactivity['band'] = request.args.get('band', None)
	basicactivity['basedir'] = satsdict[basicactivity['sat']]
	logging.warning('buildBasicActivity basicactivity {}'.format(basicactivity))
	return 0

#########################################
def traverse(basicactivity, dispatcher = None, extension = 'tif'):
	global config
	logging.warning('traverse - called traverse {}'.format(basicactivity))
	activity = copy.deepcopy(basicactivity)
	filestructure = {}
	filestructure[basicactivity['sat']] = {}
	basedir = basicactivity['basedir']

# Get the yyyy_mm, if any
	yms = []
	if 'ym' not in basicactivity:
		yms = glob.glob('{}/*'.format(basedir))
	else:
		yms = glob.glob('{}/{}*'.format(basedir,basicactivity['ym']))

# Get the drd
	drds = []
	if basicactivity['drd'] is not None:
		(satx,inst,yyyy_mm,date,year,jday) = decodeDRDFilename(basicactivity['drd'])
		drd = '{}/{}/{}'.format(basedir,yyyy_mm,basicactivity['drd'])
		logging.warning('traverse - requested drd {} '.format(drd))
		if not os.path.exists(drd):
			return filestructure
		yms = [yyyy_mm]
		drds.append(drd)

# Get the radiometric processing
	rps = []
	if basicactivity['rp'] is not None:
		rps = basicactivity['rp'].split(',')
	else:
		rps = config['publish'].keys()
	logging.warning('traverse - rps {}'.format(rps))

# For each yyyy_mm in folder structure
	for ym in sorted(yms):
		logging.warning('traverse - ym {} '.format(ym))
		bym = os.path.basename(ym)
		filestructure[basicactivity['sat']][bym] = {}
# For each drd in folder structure
		if basicactivity['drd'] is None:
			drds = glob.glob('{}/*'.format(ym))
		logging.warning('traverse - drds {} '.format(drds))
		for drd in sorted(drds):
			(satx,inst,yyyy_mm,date,year,jday) = decodeDRDFilename(drd)
			if satx is None: continue
			if basicactivity['inst'] is not None and basicactivity['inst'] != inst: continue
			if 'date' in basicactivity and basicactivity['date'] != date: continue
			if 'start_date' in basicactivity and basicactivity['start_date'] > date: continue
			if 'end_date' in basicactivity and basicactivity['end_date'] < date: continue
			bdrd = os.path.basename(drd)
			filestructure[basicactivity['sat']][bym][bdrd] = {}
			logging.warning('traverse - drd {} '.format(drd))
			activity['drd'] = drd
			activity['inst'] = inst
			activity['date'] = date
# For each path_row in folder structure
			prs = glob.glob('{}/*'.format(drd))
			for pr in sorted(prs):
				logging.warning('traverse - pr {}'.format(pr))
				bpr = os.path.basename(pr)
				bprp = bpr.split('_')
				if inst == 'HRC':
					path = bprp[0]+bprp[1]
				else:
					path = bprp[0]
				if basicactivity['path'] is not None and basicactivity['path'] != path: continue
				
				if inst == 'HRC':
					row = bprp[2]+bprp[3]
				else:
					row = bprp[1]
				if basicactivity['row'] is not None and basicactivity['row'] != row: continue
				filestructure[basicactivity['sat']][bym][bdrd][bpr] = {}
				activity['path'] = path
				activity['row'] = row
				pathrow = ''.join(bpr.split('_')[0:2])
				pathrow = path+row
				activity['pathrow'] = pathrow
# For each level_interpolation_datum in folder structure
				levdatums = glob.glob('{}/*'.format(pr))
				for levdatum in sorted(levdatums):
					logging.warning('traverse - levdatum {}'.format(levdatum))
					blevdatum = os.path.basename(levdatum)
					level = blevdatum.split('_')[0]
					if level == '1': continue
					if level == '3': continue
					if basicactivity['level'] is not None and basicactivity['level'] != level: continue
					filestructure[basicactivity['sat']][bym][bdrd][bpr][blevdatum] = {}
					activity['level'] = level
					activity['tiffdir'] = levdatum
# For each kind of product (DN or SR) specified in config file or in command line
					for rp in rps:
						logging.warning('traverse - rp {}'.format(rp))
# Check if this instrument is configured to be published
						if inst not in config['publish'][rp]:
							continue
						activity['rp'] = rp
# Given the file name template for each band, find the required files in tif repository
						filestructure[basicactivity['sat']][bym][bdrd][bpr][blevdatum][rp] = {}
						filestructure[basicactivity['sat']][bym][bdrd][bpr][blevdatum][rp]['files'] = {}
						filestructure[basicactivity['sat']][bym][bdrd][bpr][blevdatum][rp]['metadata'] = {}
						activity['files'] = {}
						activity['metadata'] = {}
						activity['sceneid'] = '{}_{}{}{}'.format(activity['sat'],activity['inst'],activity['pathrow'],activity['date'].replace('-',''))
						activity['dataset'] = '{}_{}_L{}_{}'.format(activity['sat'],activity['inst'],level,rp)
# Check if png exists
						template = '{}/*.png'.format(levdatum)
						pngs = glob.glob(template)
						logging.warning('traverse - {} {} lenpngs {}'.format(activity['sceneid'],template,len(pngs)))
						if len(pngs) > 0:
							activity['pngname'] = pngs[0]
						else:
							if 'pngname' in activity: del activity['pngname']

						for band,bandtemplate in config['publish'][rp][inst].items():
							template = '{}/*{}'.format(levdatum,bandtemplate)
							if extension == 'zip':
								template += '.'+extension
							tiffs = glob.glob(template)
							logging.warning('traverse - {} {} tiffs {}'.format(band,template,len(tiffs)))
			
							if len(tiffs) == 1:
								activity['files'][band] = filestructure[basicactivity['sat']][bym][bdrd][bpr][blevdatum][rp]['files'][band] = tiffs[0]
								xmlfile = filestructure[basicactivity['sat']][bym][bdrd][bpr][blevdatum][rp]['metadata'][band] = tiffs[0].replace('tif','xml')
								if os.path.exists(xmlfile):
									activity['metadata'][band] = xmlfile
							elif len(tiffs) > 1:
# WFI may have more than 1 file ('LEFT', 'RIGHT' and '') per band. Get the ''
								for wfifile in tiffs:
									if wfifile.find('LEFT') == -1 and wfifile.find('RIGHT') == -1:
										activity['files'][band] = filestructure[basicactivity['sat']][bym][bdrd][bpr][blevdatum][rp]['files'][band] = wfifile
										xmlfile = filestructure[basicactivity['sat']][bym][bdrd][bpr][blevdatum][rp]['metadata'][band] = tiffs[0].replace('tif','xml')
										if os.path.exists(xmlfile):
											activity['metadata'][band] = xmlfile

# if no file was found, log error and continue
						if len(filestructure[basicactivity['sat']][bym][bdrd][bpr][blevdatum][rp]['files']) == 0:
							del filestructure[basicactivity['sat']][bym][bdrd][bpr][blevdatum][rp]							
							logging.warning('traverse - no {} tiffs in {}'.format(rp,levdatum))
							continue
						if dispatcher is not None:
							logging.warning('traverse - calling dispatcher {}'.format(activity))
							dispatcher(activity,basicactivity)
	return filestructure

#########################################
def xupload(activity,basicactivity):
# If upload was requested and product is SR, upload it. For WPM the upload will not be done yet
	logging.warning('xupload - called xupload {}'.format(activity))
	inst = activity['inst']
	rp = activity['rp']
	if rp == 'SR' and inst != 'WPM':
		activity['task'] = 'uploadScene'
		db_launch(activity)
		logging.warning('xupload - calling taskschain {}'.format(activity))

		taskschain = chain([tasks.uploadScene.s(activity)])
		taskschain.apply_async()
		basicactivity['total'] += 1
		basicactivity['started'] += 1

	logging.warning('xupload - returning xupload {}'.format(activity))
	return

#########################################
@app.route('/upload', methods=['GET'])
def upload():
	basicactivity = {}
	retcode = buildBasicActivity(basicactivity)
	if retcode != 0:
		return json.dumps(basicactivity, indent=2)
	basicactivity['started'] = 0
	basicactivity['missing'] = 0
	basicactivity['total'] = 0
	basicactivity['done'] = 0

	step_start = time.time()
	basicactivity['start'] = str(time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(step_start)))
	logging.warning('upload - calling traverse {}'.format(basicactivity))
	filestructure = traverse(basicactivity,xupload)
	step_end = time.time()
	elapsedtime = step_end - step_start
	basicactivity['end'] = str(time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(step_end)))
	basicactivity['elapsed'] = str(datetime.timedelta(seconds=elapsedtime))
	return json.dumps(basicactivity, indent=2)

#########################################
def xpositioning(activity,basicactivity):
# Create the task for this scene
	activity['task'] = 'positioningScene'
	activity['sceneid'] = activity['sceneid']+'-'+activity['band']
	logging.warning('xpositioning - {}'.format(activity))
	db_launch(activity)
	if activity['band'] not in activity['files']:
		activity['message'] = 'Band {} file missing for scene in {}'.format(activity['band'],activity['tiffdir'])
		db_error(activity)
		basicactivity['missing'] += 1
		basicactivity['total'] += 1
		return
	taskschain = chain([tasks.positioningScene.s(activity)])
	taskschain.apply_async()
	basicactivity['started'] += 1

#########################################
def updatePositioningActivity(basicactivity):
	basicactivity['band'] = request.args.get('band', 'nir')
	basicactivity['savekernels'] = request.args.get('savekernels', None)
	basicactivity['savememory'] = request.args.get('savememory', None)
	basicactivity['correlation'] = float(request.args.get('correlation', 0.6))
	basicactivity['sigma'] = float(request.args.get('sigma', 3.))
	basicactivity['searchmargin'] = int(request.args.get('searchmargin', 200))
	basicactivity['resampling'] = int(request.args.get('resampling', 1))
	basicactivity['canny'] = request.args.get('canny', None)

#########################################
@app.route('/positioning', methods=['GET'])
def positioning():
	basicactivity = {}
	retcode = buildBasicActivity(basicactivity)
	if retcode != 0:
		return json.dumps(basicactivity, indent=2)
	updatePositioningActivity(basicactivity)
	basicactivity['started'] = 0
	basicactivity['missing'] = 0
	basicactivity['total'] = 0
	filestructure = traverse(basicactivity,xpositioning)
	return json.dumps(basicactivity, indent=2)

#########################################
def updateRegisteringActivity(basicactivity):
	basicactivity['band'] = request.args.get('band', 'nir')
	basicactivity['correlation'] = float(request.args.get('correlation', 0.6))
	basicactivity['searchmargin'] = int(request.args.get('searchmargin', 200))
	basicactivity['canny'] = request.args.get('canny', None)


#########################################
@app.route('/registering', methods=['GET'])
def registering():
	basicactivity = {}
	retcode = buildBasicActivity(basicactivity)
	if retcode != 0:
		return json.dumps(basicactivity, indent=2)
	updateRegisteringActivity(basicactivity)
	basicactivity['started'] = 0
	filestructure = traverse(basicactivity)
	scenes = {}
	logging.warning('registering - filestructure\n'+json.dumps(filestructure, indent=2))
	for sat in filestructure:
		logging.warning('registering - sat {}'.format(sat))
		for ym in filestructure[sat]:
			logging.warning('registering - ym {}'.format(ym))
			for drd in filestructure[sat][ym]:
				(satx,inst,yyyy_mm,date,year,jday) = decodeDRDFilename(drd)
				logging.warning('registering - drd {}'.format(drd))
				for bpr in filestructure[sat][ym][drd]:
					pathrow = ''.join(bpr.split('_')[0:2])
					logging.warning('registering - pathrow {}'.format(pathrow))
					for levdatum in filestructure[sat][ym][drd][bpr]:
						level = levdatum[0:1]
						logging.warning('registering - level {}'.format(level))
						for rp in filestructure[sat][ym][drd][bpr][levdatum]:
							dataset = '{}_{}_L{}_{}'.format(sat,inst,level,rp)
							key = dataset+'_'+pathrow
							if key not in scenes:
								scenes[key] = []
							sceneid = '{}_{}{}{}'.format(sat,inst,pathrow,date.replace('-',''))
							scene = {'sceneid':sceneid,'dataset':dataset,'rp':rp, \
							'files':filestructure[sat][ym][drd][bpr][levdatum][rp]['files'], \
							'metadata':filestructure[sat][ym][drd][bpr][levdatum][rp]['metadata'] \
							}
							logging.warning('registering - rp {}'.format(rp))
							#logging.warning('registering - append {}'.format(filestructure[sat][ym][[drd]][levdatum][rp]))
							scenes[key].append(scene)
# Group the scenes by pair
	for key in scenes:
		nscenes = len(scenes[key])
		if nscenes < 2: continue
		for idx1 in range(nscenes-1):
			idx2 = idx1 + 1
			scene1 = scenes[key][idx1]
			scene2 = scenes[key][idx2]
# Build the activity
			activity = copy.deepcopy(basicactivity)
			activity['task'] = 'registeringScene'
			activity['sceneid'] = scene1['sceneid']+'-'+scene2['sceneid']
			activity['dataset'] = scene1['dataset']
			activity['scene1'] = scene1
			activity['scene2'] = scene2
			db_launch(activity)
			mytask = tasks.registeringScene.s(copy.deepcopy(activity))	
			mytask.apply_async()

	logging.warning('registering - scenes\n'+json.dumps(scenes, indent=2))
	return json.dumps(scenes, indent=2)

#########################################
def xpublish(activity,basicactivity):
	global config
	logging.warning('xpublish - being called activity {}'.format(activity))
	sat = activity['sat']
	inst = activity['inst']
	rp = activity['rp']
	level = activity['level']

# Create the tasks list for this scene
	mytasks = []
	myfirsttask = True

# If positioning was requested
	if activity['positioning'] is not None:
		activity['task'] = 'positioningScene'
		db_launch(activity)
		if activity['band'] is not None and activity['band'] not in activity['files']:
			activity['message'] = 'Band {} file missing for scene in {}'.format(activity['band'],activity['tiffdir'])
			db_error(activity)
			basicactivity['missing'] += 1
			basicactivity['total'] += 1
			return
		
		mytask = tasks.positioningScene.s(copy.deepcopy(activity))	
		mytask.apply_async()

# Check if there is a quality band
	qualitymissing = True if 'quality' not in activity['files'] else False

# Check if this scene/dataset has already been registered
	if inst == 'XXXWPM':
		sql = "SELECT * FROM Product WHERE SceneId LIKE '{0}%%' AND Dataset = '{1}'".format(activity['sceneid'],activity['dataset'])
		presult = db_fetchall(sql)
		logging.warning('xpublish {}/{} Products {}'.format(activity['sceneid'],activity['dataset'],len(presult)))
		if len(presult) == 5*len(activity['files']):
			basicactivity['done'] += 1
			basicactivity['total'] += 1
			return
	else:
		"""
		sql = "SELECT * FROM Scene s, Product p WHERE s.SceneId = '{0}' AND p.SceneId = '{0}' AND p.Dataset = '{1}'".format(activity['sceneid'],activity['dataset'])
		presult = db_fetchall(sql)
		logging.warning('xpublish {}/{} Products {}'.format(activity['sceneid'],activity['dataset'],len(presult)))
		if len(presult) == len(activity['files']):
		"""
		sql = "SELECT * FROM Scene s, Asset p WHERE s.SceneId = '{0}' AND p.SceneId = '{0}' AND p.Dataset = '{1}'".format(activity['sceneid'],activity['dataset'])
		presult = db_fetchone(sql)
		logging.warning('xpublish  sql {} - {}'.format(sql,presult))
		if presult is not None:
			logging.warning('xpublish {}/{} Products {}'.format(activity['sceneid'],activity['dataset'],len(presult['Assets'])))
			assets = json.loads(presult['Assets'])
			pngname = presult['thumbnail']
			logging.warning('xpublish {}/{} thumbnail {}'.format(activity['sceneid'],activity['dataset'],presult['thumbnail']))
			if len(assets) == len(activity['files']) and os.path.exists(pngname):
				basicactivity['done'] += 1
				basicactivity['total'] += 1
				return

# Lets start publishing
	activity['task'] = 'publish'
	activity['message'] = ''
	activity['status'] = ''
	
# If the exact number of files was not found, log error and continue
	db_launch(activity)
	if rp == 'SR':
# quality file may not be missing
		if qualitymissing:
			activity['message'] = 'Quality file missing for scene in {}'.format(activity['tiffdir'])
			db_error(activity)
			basicactivity['missing'] += 1
			basicactivity['total'] += 1
			return
# Besides NDVI and EVI, files may not be missing
		if len(activity['files']) < len(config['publish'][rp][inst])-2:
			activity['message'] = 'Not all files found {} < {}'.format(len(activity['files']),len(config['publish'][rp][inst])-2)
			db_error(activity)
			basicactivity['missing'] += 1
			basicactivity['total'] += 1
			return
# Check when rp is DN
	else:
		if len(activity['files']) != len(config['publish'][rp][inst]):
			activity['message'] = 'Not all files found {} != {}'.format(len(activity['files']), len(config['publish'][rp][inst]))
			db_error(activity)
			basicactivity['missing'] += 1
			basicactivity['total'] += 1
			return
	db_end(activity)
# Filling the tasks list
# If SR, generate the EVI and NDVI bands
	if rp == 'SR':
		activity['task'] = 'generateVI'
		db_launch(activity)
		mytask = tasks.generateVI.s(copy.deepcopy(activity)) if myfirsttask else tasks.generateVI.s()	
		mytasks.append(mytask)
		myfirsttask = False
# If WPM, generate the the tiles. generateTiles task will launch generateQL and publishOneMS3 for each tile
	if inst == 'XXXWPM':
		activity['task'] = 'generateTiles'
		db_launch(activity)
		mytask = tasks.generateTiles.s(copy.deepcopy(activity)) if myfirsttask else tasks.generateTiles.s()	
		mytasks.append(mytask)
		myfirsttask = False
	else:
		if 'pngname' not in activity:
			activity['task'] = 'generateQL'
			db_launch(activity)
			mytask = tasks.generateQL.s(copy.deepcopy(activity)) if myfirsttask else tasks.generateQL.s()	
			mytasks.append(mytask)
			myfirsttask = False
		activity['task'] = 'publishOneMS3'
		db_launch(activity)
		mytask = tasks.publishOneMS3.s(copy.deepcopy(activity)) if myfirsttask else tasks.publishOneMS3.s()	
		mytasks.append(mytask)
		myfirsttask = False
# If upload was requested and product is SR, upload it. For WPM the upload will not be done yet
	if activity['upload'] is not None and rp == 'SR' and inst != 'WPM':
		activity['task'] = 'uploadScene'
		db_launch(activity)
		mytask = tasks.uploadScene.s()	
		mytasks.append(mytask)
		myfirsttask = False
	basicactivity['total'] += 1
	basicactivity['started'] += 1
	if activity['drd'] not in basicactivity['cq']:
		basicactivity['cq'].append(activity['drd'])
	logging.warning('xpublish - calling taskschain {}'.format(basicactivity))
	if len(mytasks) > 0:
		taskschain = chain(mytasks)
		taskschain.apply_async()
	return

#########################################
@app.route('/publish', methods=['GET'])
def publish():
	basicactivity = {}
	retcode = buildBasicActivity(basicactivity)
	if retcode != 0:
		return json.dumps(basicactivity, indent=2)
	basicactivity['upload'] = request.args.get('upload', None)
	basicactivity['positioning'] = request.args.get('positioning', None)
	if basicactivity['positioning'] is not None:
		updatePositioningActivity(basicactivity)
	basicactivity['started'] = 0
	basicactivity['missing'] = 0
	basicactivity['total'] = 0
	basicactivity['done'] = 0
	basicactivity['cq'] = []

	step_start = time.time()
	basicactivity['start'] = str(time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(step_start)))
	logging.warning('publish - calling traverse {}'.format(basicactivity))
	filestructure = traverse(basicactivity,xpublish)
	
# https://www.tutorialspoint.com/python/python_sending_email.htm
	cqnotify = os.environ.get('CQ_NOTIFY')
	if cqnotify == 'yes' and len(basicactivity['cq']) > 0:
		sender = os.environ.get('smtp_USER')
		receivers = os.environ.get('CQ_RECEIVERS').split(',')

		message = """From: %s
To: %s
Subject: Controlar passagens

Controlar passagens:
""" % (sender, ', '.join(receivers))
		logging.warning('publish - sendmail {}'.format(message))
		for drd in basicactivity['cq']:
			message = message + os.path.basename(drd)+'\n'
			logging.warning('publish - sendmail {}'.format(message))
		try:
		   smtpObj = smtplib.SMTP('smtp1.inpe.br',587)
		   smtpObj.ehlo()
		   smtpObj.starttls()
		   smtpObj.login(sender, os.environ.get('smtp_PASS'))
		   smtpObj.sendmail(sender, receivers, message)
		   smtpObj.quit()    
		   logging.warning('publish - Successfully sent email')
		   basicactivity['message'] = "Successfully sent email"
		except Exception as e:
		   logging.warning('publish - Error: unable to send email {}'.format(e))
		   basicactivity['message'] = "Error: unable to send email {}".format(e)

	step_end = time.time()
	elapsedtime = step_end - step_start
	basicactivity['end'] = str(time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(step_end)))
	basicactivity['elapsed'] = str(datetime.timedelta(seconds=elapsedtime))
	return json.dumps(basicactivity, indent=2)


#########################################
def xunzip(activity,basicactivity):
	global config
# Create the task for this scene
	activity['task'] = 'unzipScene'
	activity['message'] = ''
	inst = activity['inst']
	rp = activity['rp']
	logging.warning('xunzip - {}'.format(activity))

# Check if this task has been already done
	count = 0
	allexist = True
	for band in list(activity['files']):
		zfile = activity['files'][band]
		tfile = zfile.replace('Level-2','TIFF').replace('.zip','')
		count += 1
		if not os.path.exists(tfile):
			logging.warning('xunzip - {}/{} - {} -> {} not exist'.format(count,len(activity['files']),zfile,tfile))
			allexist = False
		else:
			logging.warning('xunzip - {}/{} - {} -> {} exist'.format(count,len(activity['files']),zfile,tfile))
	"""
	sql = "SELECT * FROM Scene s, Product p WHERE s.SceneId = '{0}' AND p.SceneId = '{0}' AND p.Dataset = '{1}'".format(activity['sceneid'],activity['dataset'])
	presult = db_fetchall(sql)
	logging.warning('xunzip {}/{} Products {}'.format(activity['sceneid'],activity['dataset'],len(presult)))
	"""
	sql = "SELECT * FROM Scene s, Asset p WHERE s.SceneId = '{0}' AND p.SceneId = '{0}' AND p.Dataset = '{1}'".format(activity['sceneid'],activity['dataset'])
	presult = db_fetchone(sql)
	if presult is not None:
		logging.warning('xunzip {}/{} Products {}'.format(activity['sceneid'],activity['dataset'],len(presult['Assets'])))
		assets = json.loads(presult['Assets'])
		if len(assets) == len(activity['files']) and allexist:
			basicactivity['done'] += 1
			basicactivity['total'] += 1
			return

	db_launch(activity)
# Check if number of zips correspond to number of bands
	if len(activity['files']) != len(config['publish'][rp][inst]):
		activity['message'] = 'Not all files found {} != {}'.format(len(activity['files']), len(config['publish'][rp][inst]))
		db_error(activity)
		basicactivity['missing'] += 1
		basicactivity['total'] += 1
		return

	taskschain = chain([tasks.unzipScene.s(activity)])
	taskschain.apply_async()
	basicactivity['started'] += 1

#########################################
@app.route('/unzip', methods=['GET'])
def unzip():
	basicactivity = {}
	retcode = buildBasicActivity(basicactivity,'/Level-2')
	if retcode != 0:
		return json.dumps(basicactivity, indent=2)
	basicactivity['started'] = 0
	basicactivity['missing'] = 0
	basicactivity['total'] = 0
	basicactivity['done'] = 0

	step_start = time.time()
	basicactivity['start'] = str(time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(step_start)))
	logging.warning('unzip - calling traverse {}'.format(basicactivity))
	filestructure = traverse(basicactivity,xunzip,'zip')
	step_end = time.time()
	elapsedtime = step_end - step_start
	basicactivity['end'] = str(time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(step_end)))
	basicactivity['elapsed'] = str(datetime.timedelta(seconds=elapsedtime))
	return json.dumps(basicactivity, indent=2)

#########################################
@app.route('/create', methods=['GET'])
def create():
	basedir = '/operation'
	file = 'template.xlsx'
	file = os.path.join(basedir,file)
	if not os.path.exists(file):
		return 'Error - File not found {}.\n'.format(request.args.get('file', None))
	wb = openpyxl.load_workbook(file)
	sheet = wb['new']
	ri = 1
	ci = 1

	sdict = {}
	while sheet.cell(row=ri, column=ci).value is not None:
		sdict[ci] = {}
		name = sheet.cell(row=ri, column=ci).value.replace(' ','').replace('-','')
		sdict[ci]['name'] = name
		type = sheet.cell(row=ri+1, column=ci).value
		sdict[ci]['type'] = type
		if sheet.cell(row=ri+2, column=ci).value is not None:
			size = sheet.cell(row=ri+2, column=ci).value
			sdict[ci]['size'] = size
		if sheet.cell(row=ri+3, column=ci).value is None or sheet.cell(row=ri+3, column=ci).value == '':
			sdict[ci]['null'] = 'NULL'
		else:
			sdict[ci]['null'] = 'NOT NULL'
		logging.warning('Columns - {} {} {}'.format(ci,sheet.cell(row=ri+3, column=ci).value,sdict[ci]))
		ci += 1

	myengine = getEngine('operation')
	if myengine is None:
		return 'Error connecting to db\n'
	sql = "DROP TABLE IF EXISTS satop"
	logging.warning(sql)
	result = myengine.execute(sql)
	sql = "CREATE TABLE IF NOT EXISTS satop (id bigint(20) NOT NULL AUTO_INCREMENT,"
	for ci in sdict.keys():
		if 'size' in sdict[ci]:
			sql += "{} {}({}) {},".format(sdict[ci]['name'],sdict[ci]['type'],sdict[ci]['size'],sdict[ci]['null'])
		else:
			sql += "{} {} {},".format(sdict[ci]['name'],sdict[ci]['type'],sdict[ci]['null'])
	sql = sql[:-1] + ",CONSTRAINT satop_pk PRIMARY KEY (id)) ENGINE=InnoDB DEFAULT CHARSET=latin1"
	logging.warning(sql)
	result = myengine.execute(sql)
	myengine.dispose()
	return 'Done - {}\n'.format(sql)

#########################################
@app.route('/load', methods=['GET'])
def load():
	basedir = '/operation'
	file = request.args.get('file', None)
	if file is None:
		return 'Error - Please enter a file (file).\n'
	file = os.path.join(basedir,file)
	if not os.path.exists(file):
		return 'Error - File not found {}.\n'.format(request.args.get('file', None))
	wb = openpyxl.load_workbook(file)
	logging.warning ('sheetnames',wb.sheetnames)
	sheet = request.args.get('sheet', 'new')
	if sheet not in wb.sheetnames:
		return 'Error - No such sheet {}\n'.format(sheet)
	sheet = wb[sheet]
	ri = 1
	ci = 1

	sdict = {}
	while sheet.cell(row=ri, column=ci).value is not None:
		sdict[ci] = sheet.cell(row=ri, column=ci).value.replace(' ','').replace('-','')
		ci += 1
	logging.warning('Number of rows - ',sheet.max_row)
	logging.warning('Columns - {}'.format(sdict))

	myengine = getEngine('operation')
	if myengine is None:
		return 'Error connecting to db\n'

	for ri in range(2,sheet.max_row+1):
		params = ''
		values = ''
		for ci in sdict.keys():
			if sheet.cell(row=ri, column=ci).value is None or sheet.cell(row=ri, column=ci).value == '-': continue
			key = sdict[ci]
			value = sheet.cell(row=ri, column=ci).value
			params += key+','
			if type(value) is str:
					values += "'{0}',".format(value)
			else:
					values += "{0},".format(value)
		
		sql = "INSERT INTO satop ({0}) VALUES({1})".format(params[:-1],values[:-1])

		logging.warning('sql',sql)
		result = myengine.execute(sql)
	myengine.dispose()
	sdict['Number of rows'] = sheet.max_row
	return 'Done - {}\n'.format(sdict)

#########################################
@app.route('/testinsert', methods=['GET'])
def testinsert():
	recs = request.args.get('recs', 500000)

	myengine = getEngine('operation')
	if myengine is None:
		return 'Error connecting to db\n'

	sql = "TRUNCATE testinsertc"
	result = myengine.execute(sql)
	sql = "TRUNCATE testinserti"
	result = myengine.execute(sql)
	for ri in range(recs):
		sql = "INSERT INTO testinsertc (id,c1,c2) VALUES('ABC{0}XYZ','{0}','{0}')".format(ri)
		result = myengine.execute(sql)
		sql = "INSERT INTO testinserti (c1,c2) VALUES('{0}','{0}')".format(ri)
		result = myengine.execute(sql)
	myengine.dispose()
	return 'Done - {}\n'.format(ri)

@app.errorhandler(400)
def handle_bad_request(e):
    resp = jsonify({'code': 400, 'message': 'Bad Request - {}'.format(e.description)})
    resp.status_code = 400
    resp.headers.add('Access-Control-Allow-Origin', '*')
    return resp


@app.errorhandler(404)
def handle_page_not_found(e):
    resp = jsonify({'code': 404, 'message': 'Page not found'})
    resp.status_code = 404
    resp.headers.add('Access-Control-Allow-Origin', '*')
    return resp


@app.errorhandler(500)
def handle_api_error(e):
    resp = jsonify({'code': 500, 'message': 'Internal Server Error'})
    resp.status_code = 500
    resp.headers.add('Access-Control-Allow-Origin', '*')
    return resp


@app.errorhandler(502)
def handle_bad_gateway_error(e):
    resp = jsonify({'code': 502, 'message': 'Bad Gateway'})
    resp.status_code = 502
    resp.headers.add('Access-Control-Allow-Origin', '*')
    return resp


@app.errorhandler(503)
def handle_service_unavailable_error(e):
    resp = jsonify({'code': 503, 'message': 'Service Unavailable'})
    resp.status_code = 503
    resp.headers.add('Access-Control-Allow-Origin', '*')
    return resp


@app.errorhandler(Exception)
def handle_exception(e):
    logging.exception(e)
    resp = jsonify({'code': 500, 'message': 'Internal Server Error'})
    resp.status_code = 500
    resp.headers.add('Access-Control-Allow-Origin', '*')
    return resp
